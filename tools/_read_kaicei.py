import openpyxl

wb = openpyxl.load_workbook(r'tools/documents/事務局立て替え分の経費.xlsx', data_only=True)
print('シート一覧:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    used = [(c.coordinate, c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    print(f'\n=== {name} (計{len(used)}セル) ===')
    for coord, val in used:
        print(f'  {coord}: {repr(val)}')
